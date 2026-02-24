"""
Extract All Sheets from Excel

Detects all sheets in an Excel file and converts each sheet using
direct function calls with parallel processing.

Usage:
    python extract_all_sheets.py <excel_file> <output_dir>

Example:
    python extract_all_sheets.py input.xlsx output/sheets/

Output:
    output/sheets/
    ├── 0.csv
    ├── 1.csv
    ├── 5.csv
    ├── 5.1.1a.csv
    ├── 5.1.1b.csv
    └── ...
    output/images/
    ├── 5.1.1a_B5_image1.png
    └── ...
"""

import sys
import os
import re
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl import load_workbook
from sheet_converter import convert_file


def sanitize_filename(sheet_name: str) -> str:
    """
    Convert sheet name to safe filename.
    
    Examples:
        '5.1.1a' -> '5.1.1a'
        '5.1.1b (2)' -> '5.1.1b_2'
        'Status' -> 'Status'
    """
    safe_name = sheet_name.replace(' ', '_')
    safe_name = safe_name.replace('(', '').replace(')', '')
    safe_name = re.sub(r'[^\w\.\-]', '_', safe_name)
    safe_name = re.sub(r'_+', '_', safe_name)
    safe_name = safe_name.strip('_')
    return safe_name


def get_sheet_names(excel_path: str) -> list:
    """Get all sheet names from Excel file."""
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def _extract_single_sheet(excel_path: str, sheet_name: str, output_file: str,
                          index: int, total: int) -> dict:
    """Extract a single sheet — worker function for parallel execution."""
    safe_name = os.path.basename(output_file)
    try:
        convert_file(excel_path, output_file, sheet_name, quiet=True)
        file_size = os.path.getsize(output_file) if os.path.exists(output_file) else 0
        print(f"[{index}/{total}] ✓ '{sheet_name}' ({file_size:,} bytes)")
        return {'status': 'success', 'sheet_name': sheet_name,
                'filename': safe_name, 'path': output_file, 'size': file_size}
    except Exception as e:
        print(f"[{index}/{total}] ✗ '{sheet_name}' - {e}")
        return {'status': 'failed', 'sheet_name': sheet_name, 'error': str(e)}


def extract_all_sheets(excel_path: str, output_dir: str,
                       max_workers: int = 10) -> dict:
    """
    Extract all sheets from Excel file using parallel direct function calls.

    Args:
        excel_path: Path to Excel file
        output_dir: Directory to save CSV files
        max_workers: Number of parallel workers

    Returns:
        Dictionary with extraction results
    """
    sheets_dir = os.path.join(output_dir, 'sheets')
    os.makedirs(sheets_dir, exist_ok=True)

    sheet_names = get_sheet_names(excel_path)
    total = len(sheet_names)
    print(f"Found {total} sheets: {sheet_names}")
    print(f"Extracting with {max_workers} workers")
    print("-" * 60)

    results = {
        'success': [],
        'failed': [],
        'sheets': {}
    }

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        futures = {}
        for i, sheet_name in enumerate(sheet_names):
            safe_name = sanitize_filename(sheet_name)
            output_file = os.path.join(sheets_dir, f"{safe_name}.csv")
            future = executor.submit(
                _extract_single_sheet, excel_path, sheet_name, output_file,
                i + 1, total
            )
            futures[future] = sheet_name

        for future in as_completed(futures):
            result = future.result()
            if result['status'] == 'success':
                results['success'].append(result['sheet_name'])
                results['sheets'][result['sheet_name']] = {
                    'filename': result['filename'],
                    'path': result['path'],
                    'size': result['size']
                }
            else:
                results['failed'].append(result['sheet_name'])
    
    # Write manifest
    manifest_path = os.path.join(output_dir, 'manifest.txt')
    with open(manifest_path, 'w', encoding='utf-8') as f:
        f.write(f"# Excel Sheet Extraction Manifest\n")
        f.write(f"# Source: {excel_path}\n")
        f.write(f"# Total sheets: {len(sheet_names)}\n")
        f.write(f"# Successful: {len(results['success'])}\n")
        f.write(f"# Failed: {len(results['failed'])}\n\n")
        
        f.write("## Sheets\n")
        for sheet_name, info in results['sheets'].items():
            f.write(f"{sheet_name}\t{info['filename']}\t{info['size']} bytes\n")
        
        if results['failed']:
            f.write("\n## Failed\n")
            for sheet_name in results['failed']:
                f.write(f"{sheet_name}\n")
    
    return results


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Extract all sheets from Excel file")
    parser.add_argument('excel_file', help='Input Excel file (.xlsx)')
    parser.add_argument('output_dir', help='Output directory')
    parser.add_argument('--workers', '-w', type=int, default=10, help='Parallel workers (default: 10)')
    args = parser.parse_args()

    if not os.path.exists(args.excel_file):
        print(f"Error: Excel file not found: {args.excel_file}")
        sys.exit(1)

    print(f"Excel file: {args.excel_file}")
    print(f"Output dir: {args.output_dir}")
    print("=" * 60)

    results = extract_all_sheets(args.excel_file, args.output_dir, args.workers)

    print("=" * 60)
    print(f"✅ Extraction complete")
    print(f"   Success: {len(results['success'])} sheets")
    print(f"   Failed:  {len(results['failed'])} sheets")
    print(f"   Output:  {args.output_dir}/sheets/")

    if results['failed']:
        print(f"\n⚠️  Failed sheets: {results['failed']}")


if __name__ == "__main__":
    main()