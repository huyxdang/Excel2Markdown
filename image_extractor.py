"""
Excel Image Extractor - Extract images from .xlsx files via ZIP/XML parsing

Extracts images from Excel worksheets by directly parsing the Office Open XML
structure inside the .xlsx ZIP archive. This bypasses openpyxl's unreliable
image loading.

Excel .xlsx structure:
    .xlsx files are ZIP archives containing:
    ├── xl/
    │   ├── worksheets/sheet1.xml    # Cell data
    │   ├── drawings/drawing1.xml    # Image anchor positions
    │   ├── drawings/_rels/*.rels    # Maps rId -> image filename
    │   └── media/image1.png         # Actual image files

Each worksheet can reference a drawing XML file. The drawing file contains
anchor elements that specify which cell an image is attached to. The .rels
file maps rId references to actual image file paths within the ZIP.
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from zipfile import ZipFile
import xml.etree.ElementTree as ET
import os
import re
from typing import Optional, Dict, List


# XML namespaces used in Office Open XML (OOXML) drawing files
NAMESPACES = {
    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    'rel': 'http://schemas.openxmlformats.org/package/2006/relationships',
}


def _get_drawing_path_from_sheet(sheet) -> Optional[str]:
    """Extract drawing XML path from an already-open worksheet.

    Used internally by sheet_converter.py to avoid an extra load_workbook call.
    """
    if hasattr(sheet, '_rels') and sheet._rels:
        for rel in sheet._rels:
            if 'drawing' in rel.Target.lower():
                drawing_path = rel.Target
                if drawing_path.startswith('..'):
                    drawing_path = 'xl' + drawing_path[2:]
                elif not drawing_path.startswith('xl/'):
                    drawing_path = 'xl/drawings/' + drawing_path.split('/')[-1]
                return drawing_path
    return None


def get_sheet_drawing_path(excel_path: str, sheet_name: str) -> Optional[str]:
    """
    Find the drawing XML path associated with a specific worksheet.

    Args:
        excel_path: Path to the .xlsx file
        sheet_name: Name of the worksheet to look up

    Returns:
        Path to the drawing XML file within the ZIP archive
        (e.g., 'xl/drawings/drawing2.xml'), or None if no drawing exists
    """
    wb = load_workbook(excel_path, data_only=False)
    sheet = wb[sheet_name]
    drawing_path = _get_drawing_path_from_sheet(sheet)
    wb.close()
    return drawing_path


def _process_anchor(
    anchor: ET.Element,
    rels: Dict[str, str],
    zf: ZipFile,
    images_dir: str,
    sheet_name: str
) -> Optional[Dict]:
    """
    Process a single image anchor element and extract the image file.

    Excel anchors images to cells using either:
    - twoCellAnchor: Image spans from one cell to another
    - oneCellAnchor: Image is anchored to a single cell

    Both types have a <from> element specifying the top-left cell position.

    Returns:
        Dictionary with image info if successful:
        {
            'cell': 'B6',
            'filename': '5.1.2a_B6_image2.png',
            'markdown': '![5.1.2a_B6](images/5.1.2a_B6_image2.png)',
            'description': '5.1.2a_B6'
        }
        Returns None if image extraction fails.
    """
    from_elem = anchor.find('xdr:from', NAMESPACES)
    if from_elem is None:
        return None

    # Excel uses 0-indexed col/row in XML, convert to 1-indexed cell reference
    col = int(from_elem.find('xdr:col', NAMESPACES).text)
    row = int(from_elem.find('xdr:row', NAMESPACES).text)
    cell_ref = f"{get_column_letter(col + 1)}{row + 1}"

    pic = anchor.find('.//xdr:pic', NAMESPACES)
    if pic is None:
        return None

    blip = pic.find('.//a:blip', NAMESPACES)
    if blip is None:
        return None

    rid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
    if rid not in rels:
        return None

    source_path = rels[rid]

    # Create unique filename: {sheet}_{cell}_{original_image_name}
    safe_sheet = re.sub(r'[^\w\-]', '_', sheet_name)
    original_img_name = os.path.basename(source_path)
    output_filename = f"{safe_sheet}_{cell_ref}_{original_img_name}"
    output_path = os.path.join(images_dir, output_filename)

    if source_path in zf.namelist():
        with open(output_path, 'wb') as f:
            f.write(zf.read(source_path))
    else:
        return None

    description = f"{sheet_name}_{cell_ref}"

    return {
        'cell': cell_ref,
        'filename': output_filename,
        'markdown': f"![{description}](images/{output_filename})",
        'description': description
    }


def extract_images_from_drawing(
    excel_path: str,
    drawing_path: str,
    output_dir: str,
    sheet_name: str
) -> List[Dict]:
    """
    Extract all images from a drawing XML file.

    Parses the drawing XML to find all image anchors (twoCellAnchor and
    oneCellAnchor elements), extracts the referenced images, and saves
    them to the output directory.

    Args:
        excel_path: Path to the .xlsx file
        drawing_path: Path to the drawing XML within the ZIP
        output_dir: Base directory for output (images saved to output_dir/images/)
        sheet_name: Name of the worksheet (used for unique filenames)

    Returns:
        List of image info dictionaries (see _process_anchor for structure)
    """
    images = []

    if not drawing_path:
        return images

    drawing_filename = drawing_path.split('/')[-1]
    rels_path = f"xl/drawings/_rels/{drawing_filename}.rels"

    with ZipFile(excel_path, 'r') as zf:
        # Build rId -> image path mapping
        rels = {}
        if rels_path in zf.namelist():
            rels_xml = zf.read(rels_path).decode('utf-8')
            rels_root = ET.fromstring(rels_xml)
            for rel in rels_root.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                rid = rel.get('Id')
                target = rel.get('Target')
                if target.startswith('..'):
                    target = 'xl' + target[2:]
                rels[rid] = target

        if not rels:
            return images

        if drawing_path not in zf.namelist():
            return images

        drawing_xml = zf.read(drawing_path).decode('utf-8')
        root = ET.fromstring(drawing_xml)

        images_dir = os.path.join(output_dir, 'images')
        os.makedirs(images_dir, exist_ok=True)

        for anchor in root.findall('.//xdr:twoCellAnchor', NAMESPACES):
            img_info = _process_anchor(anchor, rels, zf, images_dir, sheet_name)
            if img_info:
                images.append(img_info)

        for anchor in root.findall('.//xdr:oneCellAnchor', NAMESPACES):
            img_info = _process_anchor(anchor, rels, zf, images_dir, sheet_name)
            if img_info:
                images.append(img_info)

    return images


def extract_images(excel_path: str, sheet_name: str, output_dir: str) -> List[Dict]:
    """
    Extract all images from a specific worksheet.

    High-level convenience function that combines sheet lookup and image extraction.

    Args:
        excel_path: Path to the .xlsx file
        sheet_name: Name of the worksheet
        output_dir: Base directory for output

    Returns:
        List of image info dictionaries
    """
    drawing_path = get_sheet_drawing_path(excel_path, sheet_name)
    return extract_images_from_drawing(excel_path, drawing_path, output_dir, sheet_name)
