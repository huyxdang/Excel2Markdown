"""
BRD Sheet Converter - Excel Sheet to Markdown-ready CSV

Converts Excel worksheets to CSV format with cell coordinates, markdown image
references, and hyperlink extraction. Designed for LLM processing pipelines
that need structured data with embedded image links and sheet relationships.

Image extraction is handled by image_extractor.py, which parses the .xlsx
ZIP archive's internal XML to locate and extract embedded images.

Usage:
    python sheet_converter.py <excel_file> [output_csv] [sheet_name]

Example:
    python sheet_converter.py input.xlsx output.csv "5.1.2a"

Output:
    output.csv              # CSV with markdown image references and hyperlinks
    images/                 # Extracted image files
        B6_Picture_2.png
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import sys
import os
import re
from typing import Optional, Dict, List

from image_extractor import _get_drawing_path_from_sheet, extract_images_from_drawing


def _extract_hyperlinks_from_sheet(sheet) -> Dict[str, str]:
    """Extract all hyperlinks from an already-open worksheet."""
    hyperlinks = {}

    if hasattr(sheet, '_hyperlinks') and sheet._hyperlinks:
        for hyperlink in sheet._hyperlinks:
            cell_ref = hyperlink.ref
            if hyperlink.location:
                hyperlinks[cell_ref] = f"#{hyperlink.location}"
            elif hyperlink.target:
                hyperlinks[cell_ref] = hyperlink.target

    for row in sheet.iter_rows():
        for cell in row:
            if cell.hyperlink:
                cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
                if cell_ref not in hyperlinks:
                    if cell.hyperlink.location:
                        hyperlinks[cell_ref] = f"#{cell.hyperlink.location}"
                    elif cell.hyperlink.target:
                        hyperlinks[cell_ref] = cell.hyperlink.target

    return hyperlinks


def extract_hyperlinks(excel_path: str, sheet_name: str) -> Dict[str, str]:
    """
    Extract all hyperlinks from a worksheet.
    
    Excel cells can contain hyperlinks to:
    - Other sheets within the same workbook (internal links)
    - External URLs (websites, files, etc.)
    
    This function extracts both types and returns them in a dictionary
    for easy lookup during CSV generation.
    
    Args:
        excel_path: Path to the .xlsx file
        sheet_name: Name of the worksheet
        
    Returns:
        Dictionary mapping cell reference to hyperlink target
        Example: {
            'B2': '#5.1.1a!A1',      # Internal link to sheet 5.1.1a
            'B3': '#5.1.2a!A1',      # Internal link to sheet 5.1.2a  
            'C5': 'https://...',     # External URL
        }
    """
    wb = load_workbook(excel_path, data_only=False)
    sheet = wb[sheet_name]
    hyperlinks = _extract_hyperlinks_from_sheet(sheet)
    wb.close()
    return hyperlinks


def parse_internal_link(link: str) -> Optional[Dict[str, str]]:
    """
    Parse an internal Excel hyperlink to extract sheet and cell reference.
    
    Internal Excel links use the format: #SheetName!CellRef
    Sheet names with spaces or special characters are quoted: #'Sheet Name'!A1
    
    Args:
        link: Hyperlink string like '#5.1.1a!A1' or '#SheetName!B5'
        
    Returns:
        Dictionary with parsed info:
        {
            'type': 'internal',
            'sheet': '5.1.1a',
            'cell': 'A1'
        }
        Returns None if not an internal link.
    
    Examples:
        >>> parse_internal_link("#5.1.1a!A1")
        {'type': 'internal', 'sheet': '5.1.1a', 'cell': 'A1'}
        
        >>> parse_internal_link("#'Sheet with spaces'!B5")
        {'type': 'internal', 'sheet': 'Sheet with spaces', 'cell': 'B5'}
        
        >>> parse_internal_link("https://google.com")
        None
    """
    if not link.startswith('#'):
        return None
    
    # Remove leading #
    link = link[1:]
    
    # Pattern: SheetName!CellRef or 'Sheet Name'!CellRef
    # Handle quoted sheet names (for names with spaces/special chars)
    if link.startswith("'"):
        # Quoted: 'Sheet Name'!A1
        match = re.match(r"'([^']+)'!([A-Z]+\d+)", link)
    else:
        # Unquoted: SheetName!A1
        match = re.match(r"([^!]+)!([A-Z]+\d+)", link)
    
    if match:
        return {
            'type': 'internal',
            'sheet': match.group(1),
            'cell': match.group(2)
        }
    
    return None


def get_sheet_relationships(excel_path: str) -> Dict[str, List[str]]:
    """
    Build a map of sheet relationships based on hyperlinks.
    
    Analyzes all sheets to find which sheets link to which other sheets.
    This is useful for understanding the document structure and hierarchy.
    
    Args:
        excel_path: Path to the .xlsx file
        
    Returns:
        Dictionary mapping source sheet to list of target sheets
        Example: {
            'Overview': ['5.1.1a', '5.1.2a', '5.1.3a'],
            '5.1.1a': ['5.1.1b'],
            '5.1.2a': ['5.1.2b'],
        }
    """
    wb = load_workbook(excel_path, data_only=False)
    sheet_names = wb.sheetnames

    relationships = {}

    for sn in sheet_names:
        hyperlinks = _extract_hyperlinks_from_sheet(wb[sn])
        targets = set()

        for cell_ref, link in hyperlinks.items():
            parsed = parse_internal_link(link)
            if parsed and parsed['sheet'] in sheet_names:
                targets.add(parsed['sheet'])

        if targets:
            relationships[sn] = sorted(list(targets))

    wb.close()
    return relationships


def excel_to_csv(excel_path: str, sheet_name: str = None, output_dir: str = '.') -> str:
    """
    Convert an Excel worksheet to CSV with cell coordinates, markdown image refs,
    and hyperlinks.
    
    Each cell is output as "COORD: VALUE" format. Images are referenced using
    markdown syntax. Hyperlinks are converted to markdown link format:
    - Internal links: [Text](→SheetName)
    - External links: [Text](URL)
    
    Args:
        excel_path: Path to the .xlsx file
        sheet_name: Name of sheet to convert (None = first/active sheet)
        output_dir: Directory to save extracted images
        
    Returns:
        Sparse CSV string — only non-empty cells are included per row:
            Sheet: SheetName
            A1: value1,B1: [Link Text](→OtherSheet)
            B2: value2,C2: ![Image](images/C2_Picture.png)

    Note:
        Images don't count toward Excel's max_row/max_col, so this function
        extends the iteration range to include cells containing images.
    """
    # Open once with data_only=False to extract metadata (drawing path + hyperlinks)
    wb_meta = load_workbook(excel_path, data_only=False)
    if sheet_name:
        sheet_meta = wb_meta[sheet_name]
    else:
        sheet_meta = wb_meta.active
        sheet_name = sheet_meta.title

    drawing_path = _get_drawing_path_from_sheet(sheet_meta)
    hyperlinks = _extract_hyperlinks_from_sheet(sheet_meta)
    wb_meta.close()

    # Extract images via ZIP parsing (no openpyxl needed)
    images = extract_images_from_drawing(excel_path, drawing_path, output_dir, sheet_name)
    image_dict = {img['cell']: img['markdown'] for img in images}

    # Open once with data_only=True for cell values
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb[sheet_name]
    
    # Calculate max row/col including image positions
    # (images don't contribute to sheet.max_row/max_col)
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    for img in images:
        cell_ref = img['cell']
        match = re.match(r'([A-Z]+)(\d+)', cell_ref)
        if match:
            img_col = column_index_from_string(match.group(1))
            img_row = int(match.group(2))
            max_row = max(max_row, img_row)
            max_col = max(max_col, img_col)
    
    # Build sparse CSV lines (only non-empty cells)
    csv_lines = [f"Sheet: {sheet.title}"]

    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        row_values = []
        for cell in row:
            col_letter = get_column_letter(cell.column)
            row_num = cell.row
            cell_coord = f"{col_letter}{row_num}"

            if cell_coord in image_dict:
                row_values.append(f"{cell_coord}: {image_dict[cell_coord]}")
            elif cell.value is not None:
                value = str(cell.value)
                value = value.replace('\n', '\\n')

                if cell_coord in hyperlinks:
                    link = hyperlinks[cell_coord]
                    parsed = parse_internal_link(link)

                    if parsed:
                        value = f"[{value}](→{parsed['sheet']})"
                    else:
                        value = f"[{value}]({link})"

                row_values.append(f"{cell_coord}: {value}")
            # Skip empty cells entirely (sparse format)

        if row_values:
            csv_lines.append(','.join(row_values))
    
    wb.close()
    return '\n'.join(csv_lines)


def convert_all_sheets(excel_path: str, output_path: str = None) -> str:
    """
    Convert ALL sheets in an Excel workbook to a single CSV file.
    
    Each sheet is separated by a clear delimiter for easy parsing.
    Sheet relationships (hyperlinks between sheets) are preserved.
    
    Args:
        excel_path: Path to the .xlsx file
        output_path: Path to save CSV (None = return string only)
        
    Returns:
        Combined CSV string with all sheets
        
    Output format:
        ================================================================================
        Sheet: Overview
        ================================================================================
        A1: Feature ID,B1: [Feature Name](→5.1.1a)
        ...
        
        ================================================================================
        Sheet: 5.1.1a
        ================================================================================
        A1: Detailed specs...
        ...
    """
    wb = load_workbook(excel_path, data_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    
    output_dir = os.path.dirname(output_path) or '.' if output_path else '.'
    
    all_csv_parts = []
    
    # First, show sheet relationships for context
    relationships = get_sheet_relationships(excel_path)
    if relationships:
        all_csv_parts.append("# Sheet Relationships (for reference)")
        for source, targets in relationships.items():
            all_csv_parts.append(f"# {source} → {', '.join(targets)}")
        all_csv_parts.append("")
    
    # Convert each sheet
    for sheet_name in sheet_names:
        separator = "=" * 80
        all_csv_parts.append(separator)
        all_csv_parts.append(f"Sheet: {sheet_name}")
        all_csv_parts.append(separator)
        
        csv_content = excel_to_csv(excel_path, sheet_name, output_dir)
        # Remove the "Sheet: name" line since we already added it
        lines = csv_content.split('\n')
        if lines and lines[0].startswith('Sheet:'):
            lines = lines[1:]
        all_csv_parts.append('\n'.join(lines))
        all_csv_parts.append("")  # Blank line between sheets
    
    combined_csv = '\n'.join(all_csv_parts)
    
    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(combined_csv)
        print(f"Saved CSV to: {output_path}")
        print(f"Images saved to: {os.path.join(output_dir, 'images')}/")
        print(f"Total sheets: {len(sheet_names)}")
        if relationships:
            print(f"Sheets with links: {list(relationships.keys())}")
    
    return combined_csv


def convert_file(
    excel_path: str,
    output_path: str = None,
    sheet_name: str = None,
    quiet: bool = False
) -> str:
    """
    Convert an Excel file to CSV with extracted images and hyperlinks.

    Main entry point for the converter. Handles file I/O and provides
    user feedback about output locations.

    Args:
        excel_path: Path to the .xlsx file
        output_path: Path to save output (None = print to stdout)
        sheet_name: Sheet to convert (None = ALL sheets)
        quiet: If True, suppress progress output

    Returns:
        The generated CSV string

    Side effects:
        - Creates output file at output_path (if specified)
        - Creates images/ directory with extracted images
    """
    if sheet_name is None:
        return convert_all_sheets(excel_path, output_path)

    output_dir = (os.path.dirname(output_path) or '.') if output_path else '.'
    output_string = excel_to_csv(excel_path, sheet_name, output_dir)

    if output_path:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(output_string)
        if not quiet:
            print(f"Saved CSV to: {output_path}")
            print(f"Images saved to: {os.path.join(output_dir, 'images')}/")
    else:
        print(output_string)

    return output_string


if __name__ == "__main__":
    import argparse as _argparse

    _parser = _argparse.ArgumentParser(description="Convert Excel sheet to CSV")
    _parser.add_argument('excel_file', help='Input Excel file (.xlsx)')
    _parser.add_argument('output_file', nargs='?', default=None, help='Output CSV file path')
    _parser.add_argument('sheet_name', nargs='?', default=None, help='Sheet name (omit for all sheets)')
    _args = _parser.parse_args()

    convert_file(_args.excel_file, _args.output_file, _args.sheet_name)